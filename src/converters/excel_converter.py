"""
Конвертер Excel файлов (.xlsx, .xls, .xlsm, .xlsb) в Markdown или CSV
"""
from pathlib import Path
import pandas as pd
import re
from loguru import logger

from .base import FileConverter
from .. import config


class ExcelConverter(FileConverter):
    """
    Конвертер Excel файлов в Markdown

    Использует pandas для чтения Excel и форматирования в markdown
    Каждый лист Excel -> отдельная секция в Markdown
    """

    def __init__(self, max_rows: int = None, max_columns: int = None, sheets_limit: int = None):
        """
        :param max_rows: Максимальное количество строк на лист (None = без ограничений)
        :param max_columns: Максимальное количество столбцов (None = без ограничений)
        :param sheets_limit: Максимальное количество листов для конвертации (None = без ограничений)
        """
        super().__init__(['.xlsx', '.xls', '.xlsm', '.xlsb'])
        self.max_rows = max_rows
        self.max_columns = max_columns
        self.sheets_limit = sheets_limit

    def _sanitize_sheet_name(self, sheet_name: str) -> str:
        """
        Очищает название листа от запрещенных символов для использования в имени файла

        :param sheet_name: Название листа
        :return: Безопасное название для имени файла
        """
        # Запрещенные символы в Windows: < > : " / \ | ? *
        # В Linux/Mac: / и null
        forbidden_chars = r'[<>:"/\\|?*\x00-\x1f]'

        # Заменяем запрещенные символы на подчеркивание
        safe_name = re.sub(forbidden_chars, '_', sheet_name)

        # Убираем множественные подчеркивания
        safe_name = re.sub(r'_+', '_', safe_name)

        # Убираем подчеркивания в начале и конце
        safe_name = safe_name.strip('_')

        # Ограничиваем длину (Windows имеет лимит 255 символов на путь)
        # Оставляем 50 символов для названия листа
        if len(safe_name) > 50:
            safe_name = safe_name[:50].rstrip('_')

        # Если после очистки название пустое - используем placeholder
        if not safe_name or safe_name.isspace():
            safe_name = "Лист"

        return safe_name

    def convert(self, input_path: Path, output_path: Path) -> bool:
        """
        Конвертирует Excel файл в Markdown или CSV (в зависимости от настройки EXCEL_TO_CSV)

        :param input_path: Путь к Excel файлу
        :param output_path: Путь к .md или .csv файлу
        :return: True если успешно
        """
        # Проверяем, нужно ли сохранять как CSV
        if hasattr(config, 'EXCEL_TO_CSV') and config.EXCEL_TO_CSV:
            # Изменяем расширение output_path на .csv
            output_path = output_path.with_suffix('.csv')
            return self._convert_to_csv(input_path, output_path)
        else:
            return self._convert_to_markdown(input_path, output_path)

    def _convert_to_csv(self, input_path: Path, output_path: Path) -> bool:
        """
        Конвертирует Excel файл в CSV

        :param input_path: Путь к Excel файлу
        :param output_path: Путь к .csv файлу
        :return: True если успешно
        """
        try:
            # Для .xls используем xlrd engine
            engine = 'openpyxl'
            if input_path.suffix.lower() == '.xls':
                engine = 'xlrd'
            elif input_path.suffix.lower() == '.xlsb':
                engine = 'pyxlsb'

            # Получаем список всех листов
            excel_file = None
            last_error = None

            # Попытка 1: Стандартное открытие
            try:
                excel_file = pd.ExcelFile(input_path, engine=engine)
                sheet_names = excel_file.sheet_names
            except Exception as e:
                last_error = e
                logger.debug(f"Попытка 1 не удалась ({engine}): {e}")

                # Попытка 2: Для openpyxl пробуем с data_only=True (игнорирует формулы)
                if engine == 'openpyxl':
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(input_path, data_only=True, read_only=False)
                        sheet_names = wb.sheetnames
                        wb.close()
                        # Создаем ExcelFile после успешной проверки
                        excel_file = pd.ExcelFile(input_path, engine=engine)
                        logger.info(f"Excel файл открыт с data_only=True: {input_path.name}")
                    except Exception as e2:
                        last_error = e2
                        logger.debug(f"Попытка 2 не удалась (data_only=True): {e2}")

                        # Попытка 3: read_only=True
                        try:
                            import openpyxl
                            wb = openpyxl.load_workbook(input_path, read_only=True)
                            sheet_names = wb.sheetnames
                            wb.close()
                            excel_file = pd.ExcelFile(input_path, engine=engine)
                            logger.info(f"Excel файл открыт с read_only=True: {input_path.name}")
                        except Exception as e3:
                            last_error = e3
                            logger.debug(f"Попытка 3 не удалась (read_only=True): {e3}")
                            excel_file = None

            if excel_file is None:
                logger.error(f"Не удалось открыть Excel файл {input_path.name}: {last_error}")
                logger.error(f"Файл поврежден и не может быть восстановлен автоматически")
                return False

            # Если один лист - сохраняем как есть
            if len(sheet_names) == 1:
                df = pd.read_excel(input_path, sheet_name=sheet_names[0], engine=engine)
                output_path.parent.mkdir(parents=True, exist_ok=True)
                df.to_csv(output_path, index=False, encoding='utf-8-sig')
                logger.debug(f"Excel -> CSV (1 лист): {input_path.name}")
                return True

            # Если несколько листов - сохраняем каждый в отдельный CSV
            output_path.parent.mkdir(parents=True, exist_ok=True)
            base_name = output_path.stem

            for i, sheet_name in enumerate(sheet_names):
                try:
                    df = pd.read_excel(input_path, sheet_name=sheet_name, engine=engine)

                    # Пропускаем пустые листы
                    if df.empty:
                        continue

                    # Очищаем название листа от запрещенных символов
                    safe_sheet_name = self._sanitize_sheet_name(sheet_name)

                    # Создаем имя файла: номер + название листа
                    # Например: продажи.xlsx_лист1_Январь.csv, продажи.xlsx_лист2_Февраль.csv
                    sheet_output_path = output_path.parent / f"{base_name}_лист{i+1}_{safe_sheet_name}.csv"

                    df.to_csv(sheet_output_path, index=False, encoding='utf-8-sig')
                    logger.debug(f"Excel -> CSV (лист {i+1} '{sheet_name}'): {sheet_output_path.name}")

                except Exception as e:
                    logger.error(f"Ошибка при обработке листа '{sheet_name}': {e}")
                    continue

            logger.debug(f"Excel -> CSV ({len(sheet_names)} листов): {input_path.name}")
            return True

        except Exception as e:
            logger.error(f"Ошибка конвертации Excel -> CSV {input_path.name}: {e}")
            return False

    def _convert_to_markdown(self, input_path: Path, output_path: Path) -> bool:
        """
        Конвертирует Excel файл в Markdown

        :param input_path: Путь к Excel файлу
        :param output_path: Путь к .md файлу
        :return: True если успешно
        """
        try:
            # Читаем все листы из Excel
            # Для .xls используем xlrd engine
            engine = 'openpyxl'
            if input_path.suffix.lower() == '.xls':
                engine = 'xlrd'
            elif input_path.suffix.lower() == '.xlsb':
                engine = 'pyxlsb'

            # Получаем список всех листов
            excel_file = None
            last_error = None

            # Попытка 1: Стандартное открытие
            try:
                excel_file = pd.ExcelFile(input_path, engine=engine)
                sheet_names = excel_file.sheet_names
            except Exception as e:
                last_error = e
                logger.debug(f"Попытка 1 не удалась ({engine}): {e}")

                # Попытка 2: Для openpyxl пробуем с data_only=True (игнорирует формулы)
                if engine == 'openpyxl':
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(input_path, data_only=True, read_only=False)
                        sheet_names = wb.sheetnames
                        wb.close()
                        # Создаем ExcelFile после успешной проверки
                        excel_file = pd.ExcelFile(input_path, engine=engine)
                        logger.info(f"Excel файл открыт с data_only=True: {input_path.name}")
                    except Exception as e2:
                        last_error = e2
                        logger.debug(f"Попытка 2 не удалась (data_only=True): {e2}")

                        # Попытка 3: read_only=True
                        try:
                            import openpyxl
                            wb = openpyxl.load_workbook(input_path, read_only=True)
                            sheet_names = wb.sheetnames
                            wb.close()
                            excel_file = pd.ExcelFile(input_path, engine=engine)
                            logger.info(f"Excel файл открыт с read_only=True: {input_path.name}")
                        except Exception as e3:
                            last_error = e3
                            logger.debug(f"Попытка 3 не удалась (read_only=True): {e3}")
                            excel_file = None

            if excel_file is None:
                logger.error(f"Не удалось открыть Excel файл {input_path.name}: {last_error}")
                logger.error(f"Файл поврежден и не может быть восстановлен автоматически")
                return False

            # Ограничиваем количество листов
            total_sheets = len(sheet_names)
            sheets_to_process = sheet_names if self.sheets_limit is None else sheet_names[:self.sheets_limit]
            is_truncated_sheets = self.sheets_limit is not None and total_sheets > self.sheets_limit

            # Создаем метаданные
            metadata = self._create_metadata(input_path, total_sheets, is_truncated_sheets)

            # Конвертируем каждый лист
            all_content = [metadata]

            for i, sheet_name in enumerate(sheets_to_process, 1):
                try:
                    # Читаем лист
                    df = pd.read_excel(
                        input_path,
                        sheet_name=sheet_name,
                        engine=engine
                    )

                    # Пропускаем пустые листы
                    if df.empty:
                        logger.debug(f"Пропущен пустой лист: {sheet_name}")
                        continue

                    # Получаем информацию о размере
                    total_rows, total_cols = df.shape
                    is_truncated_rows = self.max_rows is not None and total_rows > self.max_rows
                    is_truncated_cols = self.max_columns is not None and total_cols > self.max_columns

                    # Ограничиваем размер
                    df_display = df
                    if is_truncated_rows:
                        df_display = df_display.head(self.max_rows)

                    if is_truncated_cols:
                        df_display = df_display.iloc[:, :self.max_columns]

                    # Добавляем заголовок листа
                    sheet_header = f"\n## Лист {i}: {sheet_name}\n\n"
                    sheet_header += f"**Размер:** {total_rows} строк × {total_cols} столбцов\n\n"

                    # Конвертируем в markdown таблицу
                    markdown_table = df_display.to_markdown(index=False)

                    # Fallback если tabulate не установлен
                    if markdown_table is None:
                        logger.warning("tabulate не установлен, используем fallback")
                        markdown_table = self._dataframe_to_markdown_fallback(df_display)

                    # Добавляем примечание если данные обрезаны
                    truncation_note = ""
                    if is_truncated_rows or is_truncated_cols:
                        truncation_note = "\n\n"
                        if is_truncated_rows:
                            truncation_note += f"⚠️ Отображены первые {self.max_rows} из {total_rows} строк.\n\n"
                        if is_truncated_cols:
                            truncation_note += f"⚠️ Отображены первые {self.max_columns} из {total_cols} столбцов.\n\n"

                    # Собираем контент листа
                    sheet_content = sheet_header + markdown_table + truncation_note
                    all_content.append(sheet_content)

                except Exception as e:
                    logger.error(f"Ошибка при обработке листа '{sheet_name}': {e}")
                    # Продолжаем обработку других листов
                    all_content.append(f"\n## Лист {i}: {sheet_name}\n\n❌ Ошибка при конвертации листа: {str(e)}\n\n")

            # Добавляем примечание о необработанных листах
            if is_truncated_sheets:
                all_content.append(
                    f"\n---\n\n⚠️ Обработаны первые {self.sheets_limit} из {total_sheets} листов.\n"
                )

            # Сохраняем результат
            full_content = "\n".join(all_content)
            with open(output_path, 'w', encoding='utf-8') as md_file:
                md_file.write(full_content)

            return True

        except Exception as e:
            logger.error(f"Ошибка конвертации Excel {input_path.name}: {e}")
            return False

    def _create_metadata(self, input_path: Path, total_sheets: int, truncated: bool) -> str:
        """
        Создает метаданные для markdown файла

        :param input_path: Путь к исходному файлу
        :param total_sheets: Количество листов
        :param truncated: Обрезаны ли листы
        :return: Строка с метаданными
        """
        metadata = f"""---
source_file: {input_path.name}
original_format: {input_path.suffix}
total_sheets: {total_sheets}
sheets_truncated: {truncated}
converted_by: ExcelConverter
---

# {input_path.stem}

**Тип:** Excel документ
**Листов:** {total_sheets}

"""
        return metadata

    def _dataframe_to_markdown_fallback(self, df: pd.DataFrame) -> str:
        """
        Fallback метод для конвертации DataFrame в markdown

        :param df: DataFrame для конвертации
        :return: Markdown строка
        """
        # Создаем заголовок
        header = "| " + " | ".join(str(col) for col in df.columns) + " |"
        separator = "|" + "|".join(["---" for _ in df.columns]) + "|"

        # Создаем строки
        rows = []
        for _, row in df.iterrows():
            row_str = "| " + " | ".join(str(val) for val in row) + " |"
            rows.append(row_str)

        # Собираем таблицу
        table = "\n".join([header, separator] + rows)
        return table
